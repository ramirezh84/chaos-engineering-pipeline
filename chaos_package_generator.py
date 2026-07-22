#!/usr/bin/env python3
"""
chaos_package_generator.py — end-to-end chaos engineering package generator.

Ingests: journey graph JSON files (content-sniffed: top-level journey/nodes/edges)
         + per-service scan-card JSON files (top-level "service" field)
Emits:   a complete chaos engineering package (.xlsx):
         1. Executive Summary — focus areas ranked, with reasons citing evidence
         2. Scenario Catalog — generated rows across layers: Infra / App / Upstream /
            Downstream / Saga / Detection, each with hypothesis, why (evidence),
            tooling, computed Likelihood x Severity, tier, rating, wave
         3. Assertion Matrix — one injection, many journey-specific assertions
         4. Coverage Gaps — unscanned nodes, unresolved identities, unmapped vias

Design: a TEMPLATE LIBRARY expands against the graph. Templates fire only when their
evidence precondition holds (accuracy contract: no evidence -> no scenario, it lands
in Coverage Gaps instead). Scores are computed:
    Severity  = base_severity + criticality_bump + data_class_bump  (capped 1..5)
    Likelihood = base_likelihood + evidence_modifiers               (capped 1..5)
Red-graded edges emit CONFIRMED_DEFECT rows (regression guards), not hypotheses.

Usage:
  python3 chaos_package_generator.py --root <dir with journey jsons> \
      --scan-cards <dir or files with service scan cards> \
      --out chaos_package.xlsx
"""
import argparse, glob, json, os, re

# ---------------------------------------------------------------- ingestion

def load_json_safe(path):
    try:
        with open(path) as f:
            return json.load(f)
    except (json.JSONDecodeError, UnicodeDecodeError, OSError):
        return None

def is_journey(obj):
    return isinstance(obj, dict) and "journey" in obj and "edges" in obj and "nodes" in obj

def is_scan_card(obj):
    return isinstance(obj, dict) and "service" in obj and not is_journey(obj)

def discover(root_dirs):
    journeys, cards = [], {}
    for root in root_dirs:
        paths = [root] if os.path.isfile(root) else glob.glob(os.path.join(root, "**", "*.json"), recursive=True)
        for p in paths:
            obj = load_json_safe(p)
            if is_journey(obj):
                journeys.append((p, obj))
            elif is_scan_card(obj):
                cards[obj["service"]] = (p, obj)
    return journeys, cards

# ---------------------------------------------------------------- scoring

CRIT_BUMP = {"P0": 1, "P1": 0, "P2": -1, "P3": -2}
DATA_BUMP = {"money-movement": 1, "state-mutating": 0, "read-only": -1}

def clamp(v):
    return max(1, min(5, v))

def score(base_l, base_s, criticality, data_class, l_mod=0):
    s = clamp(base_s + CRIT_BUMP.get(criticality, 0) + DATA_BUMP.get(data_class, 0))
    l = clamp(base_l + l_mod)
    return l, s

# ---------------------------------------------------------------- infra-md ingestion
# Per-service "<app>-infra-components-by-env.md" files: owner-generated markdown that
# declares each service's deployed infra per environment (PROD vs PERF), pinned to
# repo SHAs. When present for a service, it SUPERSEDES components.json for that
# service: components are detected from the document, key values are extracted, and
# PROD-vs-PERF deltas become validity findings. Parsing is keyword/regex based and
# deliberately tolerant — anything not found is simply not claimed.

INFRA_MD_COMPONENT_KEYWORDS = {
    "fargate": "ecs-fargate",
    "aurora": "aurora-global",
    "elasticache": "elasticache",
    "kafka": "kafka-mbus",
    "route53": "route53",
    "route 53": "route53",
    "apinlb": "nlb",
    "appnlb": "nlb",
    "appalb": "alb",
    "apigw": "api-gateway",
    "api gw": "api-gateway",
    " s3 ": "s3",
    "| s3 |": "s3",
    "kms": "kms",
    "token secret": "secrets-manager",
    "eventbridge": "eventbridge",
    "dynamodb": "dynamodb",
    "sqs": "sqs",
}


def parse_infra_md(path):
    import re as _re
    try:
        with open(path) as f:
            text = f.read()
    except OSError:
        return None
    low = text.lower()

    # service id: "for **<service> (<ACR>)**" pattern, else filename stem
    m = _re.search(r"for \*\*([a-z0-9-]+) \(", text)
    service = m.group(1) if m else os.path.basename(path).split("-infra-components")[0]

    comps = set()
    for kw, comp in INFRA_MD_COMPONENT_KEYWORDS.items():
        if kw in low:
            comps.add(comp)

    facts = {"source_file": os.path.basename(path)}
    m = _re.search(r"min (\d+) / max (\d+)", text)
    if m:
        facts["autoscale_min"], facts["autoscale_max"] = int(m.group(1)), int(m.group(2))
    m = _re.search(r"AZs per region \| \*{0,2}(\d+)\*{0,2}[^|]*\| \*{0,2}(\d+)\*{0,2}", text)
    if m:
        facts["azs_prod"], facts["azs_perf"] = int(m.group(1)), int(m.group(2))
    m = _re.search(r"TCP idle (\d+)s", text)
    if m:
        facts["nlb_idle_s"] = int(m.group(1))
    m = _re.search(r"appalb \(idle (\d+)s\)", text)
    if m:
        facts["alb_idle_s"] = int(m.group(1))
    facts["dns_simple_cname"] = "SIMPLE CNAME" in text and "not health-based" in text
    facts["perf_cache_disabled"] = ("caching DISABLED" in text) or ("caching disabled in perf" in low)
    facts["no_circuit_breaker"] = "no circuit breaker" in low
    m = _re.search(r"Replica baseline \| (\d+) \| (\d+)", text)
    if m:
        facts["replica_baseline"] = int(m.group(1))
    return {"service": service, "components": sorted(comps), "facts": facts}


# ---------------------------------------------------------------- infra component templates
# Per-component scenario templates. The generator instantiates, for each service,
# ONLY the components that service is declared to have — either from a hand-filled
# components.json (--components) or, later, from the infra repo scan output.
# Format: component -> list of (key_suffix, scenario, hypothesis, tooling, base_L, base_S, wave_or_None)
COMPONENT_TEMPLATES = {
    "ecs-fargate": [
        ("task-kill", "Fargate task termination — replacement & drain",
         "ECS replaces the task within SLO; ALB drains cleanly; only in-flight requests affected.",
         "FIS aws:ecs:stop-task", 4, 3, None),
        ("az", "AZ impairment — multi-AZ absorption",
         "Losing one AZ leaves the service above its capacity floor; no single-AZ SPOF.",
         "FIS network blackhole / NACL (one AZ)", 2, 5, None),
    ],
    "aurora-global": [
        ("db-writer-failover", "Aurora writer failover under load",
         "Connections re-establish within the pool timeout budget; bounded error window; no lost/duplicate writes.",
         "FIS / failover-db-cluster", 3, 4, None),
        ("db-region-failover", "Aurora Global regional failover drill",
         "Cross-region promotion + app cutover meets RTO; data loss within RPO.",
         "Drill (scheduled, sign-off)", 1, 5, 4),
    ],
    "rds-proxy": [
        ("proxy-down", "RDS Proxy endpoint unavailability",
         "Service fails fast per timeout chain (no hang); recovers on restore without restart.",
         "Network deny on proxy endpoint (PERF)", 2, 4, None),
    ],
    "elasticache": [
        ("cache-failover", "ElastiCache primary failover",
         "Clients reconnect within seconds; brief blip absorbed; no correctness impact.",
         "FIS / console failover", 3, 3, None),
        ("cache-down", "Cache fully unavailable — degrade to database",
         "Service serves correct (slower) responses from the DB path; DB absorbs the amplified load.",
         "Security-group deny on Redis (PERF)", 2, 4, None),
    ],
    "kafka-mbus": [
        ("kafka-unreachable", "Kafka broker (MBUS) unreachable",
         "Producers buffer without unbounded growth; consumers idle without crash-loop; no-loss catch-up on restore.",
         "Network blackhole to MBUS (PERF)", 3, 5, None),
        ("kafka-ida", "IDP token failure (Kafka OAUTHBEARER re-auth)",
         "Sessions survive to re-auth; failed re-auth backs off sanely; failure signature distinct from broker loss.",
         "IDP path block (PERF)", 2, 4, None),
    ],
    "sqs": [
        ("sqs-dlq", "SQS processing failure — DLQ & redrive",
         "Failed messages land in the DLQ (not lost, not retried forever); redrive restores processing without duplicates.",
         "Failure injection on consumer (PERF)", 3, 4, None),
    ],
    "sns": [
        ("sns-delivery", "SNS delivery failure to a subscriber",
         "Failed deliveries retry per policy and are visible; no silent notification loss.",
         "Subscriber block (PERF)", 2, 3, None),
    ],
    "lambda": [
        ("lambda-throttle", "Lambda concurrency throttling",
         "Throttled invocations retry via their source (SQS/S3 event) without loss or duplication.",
         "Reserved-concurrency limit (PERF)", 3, 4, None),
    ],
    "eventbridge": [
        ("schedule-misfire", "Scheduler misfire / double-fire",
         "A missed window is detected; a double-fire causes no duplicate side-effects.",
         "Rule disable / duplicate trigger (PERF)", 2, 4, None),
    ],
    "s3": [
        ("s3-deny", "S3 bucket unavailable / access denied",
         "Reads/writes fail visibly and resume idempotently on restore; no silent skips.",
         "Bucket policy deny (PERF)", 2, 4, None),
    ],
    "api-gateway": [
        ("apigw-throttle", "API Gateway throttle breach",
         "Clean 429s; callers back off instead of retry-hammering.",
         "Load past configured limits", 3, 3, None),
    ],
    "nlb": [
        ("nlb-idle", "NLB idle-timeout behavior (silent drop)",
         "No legitimate call path approaches the NLB idle limit; keepalives prevent surprise resets.",
         "Long-idle connection probe", 2, 3, None),
    ],
    "alb": [
        ("alb-drain", "ALB deregistration/drain during deploy under load",
         "Zero in-flight kills: drain timeout covers p99 request duration.",
         "Deploy under load", 3, 3, None),
    ],
    "route53": [
        ("dns-ttl", "Route 53 failover + client re-resolution vs TTL",
         "Clients re-resolve within the TTL budget; no stragglers pinned to the failed region.",
         "Drill + resolution measurement", 2, 4, 4),
    ],
    "secrets-manager": [
        ("secrets-outage", "Secrets retrieval outage at task recycle",
         "Running tasks unaffected; recycled tasks fail startup fast and loud — no silent degraded serving.",
         "Deny + forced recycle (PERF)", 2, 4, None),
    ],
    "parameter-store": [
        ("param-reload", "Parameter Store reload failure",
         "Service continues on last-known-good config; the missed reload is detected; resumes without restart.",
         "Deny after startup + config push (PERF)", 2, 3, None),
    ],
    "kms": [
        ("kms-deny", "KMS key access denial",
         "Encrypt/decrypt fails closed with an explicit error — no silent drop, no plaintext fallback; resumes on restore.",
         "Key-policy deny (PERF)", 2, 5, None),
    ],
    "dynamodb": [
        ("ddb-throttle", "DynamoDB throttling",
         "Throttled operations retry with backoff; no data loss; latency degradation bounded.",
         "Lowered capacity / injected throttles (PERF)", 3, 3, None),
    ],
}


def infra_component_templates(pkg, nid, card, jf, jid, crit, j_dc, node_kafka_roles,
                              declared_components, infra_md=None):
    """Per-service infra scenarios, driven by the service's declared component list.
    Precedence: infra-md file (evidence, per-env) > components.json (hand-declared)."""
    facts = (infra_md or {}).get("facts", {})
    if infra_md:
        comps = infra_md["components"]
        source_note = (f"Declared in {facts.get('source_file')} (per-env, SHA-pinned) — "
                       f"evidence-grade component declaration.")
        status = "Proposed (evidence-confirmed — infra-md)"
        # --- env-delta findings: PROD vs PERF differences that change test validity ---
        if facts.get("perf_cache_disabled"):
            pkg.gap(jid, "ENV_DELTA_PERF_VALIDITY",
                    f"{nid}: app-side Redis caching is DISABLED in PERF but ENABLED in "
                    f"PROD ({facts.get('source_file')}). Every cache-related scenario run "
                    f"in PERF tests a DIFFERENT system than PROD — cache scenarios need a "
                    f"cache-enabled environment or an explicit validity caveat.")
        if facts.get("azs_prod") and facts.get("azs_perf") and facts["azs_prod"] != facts["azs_perf"]:
            pkg.gap(jid, "ENV_DELTA_PERF_VALIDITY",
                    f"{nid}: PROD has {facts['azs_prod']} AZs per region, PERF has "
                    f"{facts['azs_perf']} ({facts.get('source_file')}). AZ-impairment "
                    f"results from PERF (N-1 of {facts['azs_perf']}) do not directly "
                    f"transfer to PROD (N-1 of {facts['azs_prod']}).")
        # --- conflicting-evidence check: journey survivability vs deploy-repo autoscaling ---
        for n in jf.get("nodes", []):
            if n.get("id") == nid:
                rc = ((n.get("survivability") or {}).get("compute") or {}).get("replica_count")
                if rc == 1 and facts.get("autoscale_min", 0) >= 2:
                    pkg.gap(jid, "CONFLICTING_EVIDENCE",
                            f"{nid}: journey file says replica_count=1, but "
                            f"{facts.get('source_file')} says autoscaling min "
                            f"{facts['autoscale_min']} / max {facts.get('autoscale_max')} "
                            f"(replica baseline {facts.get('replica_baseline')}). These "
                            f"cannot both describe steady state — reconcile before "
                            f"trusting the single-replica-exposure scenario. Likely: "
                            f"baseline 1 at deploy, autoscaling floor raises to "
                            f"{facts['autoscale_min']} — CONFIRM, do not assume.")
                    # also flag the affected scenario row itself, if it was generated
                    sr = pkg.rows.get(("infra-single-replica", nid))
                    if sr:
                        sr["status"] = "NEEDS_ASSESSMENT — conflicting replica evidence"
                        sr["why"] += (f" CONFLICT: {facts.get('source_file')} says "
                                      f"autoscaling min {facts['autoscale_min']} — "
                                      f"reconcile before running.")
    elif declared_components.get(nid):
        comps = declared_components.get(nid)
        source_note = ("Declared in components.json (owner-provided) — will be confirmed/"
                       "replaced by the infra-md file or infra repo scan.")
        status = "DECLARED — confirm with infra scan"
    else:
        # no declaration for this service: generate nothing component-specific,
        # make the gap loud instead of guessing.
        pkg.gap(jid, "NO_COMPONENT_DECLARATION",
                f"{nid} — no infra-md file, not in components.json, no infra scan output; "
                f"component-level infra scenarios NOT generated. Provide "
                f"<app>-infra-components-by-env.md (preferred) or add to components.json.")
        comps = []
        source_note = ""
        status = ""

    # Where a richer evidence-based template already covers the same fault (the
    # survivability templates fire first, with real replica/region values), merge the
    # generic component row into it instead of duplicating.
    DEDUP_ALIAS = {
        "infra-ecs-fargate-task-kill": "infra-single-replica",
        "infra-aurora-global-db-writer-failover": "infra-db-failover",
        "infra-aurora-global-db-region-failover": "infra-region-failover",
    }
    for comp in comps:
        templates = COMPONENT_TEMPLATES.get(comp)
        if templates is None:
            pkg.gap(jid, "UNKNOWN_COMPONENT",
                    f"{nid} declares component '{comp}' which has no template yet — add it "
                    f"to COMPONENT_TEMPLATES rather than ignoring it.")
            continue
        for key, scen, hypo, tool, bl, bs, wave in templates:
            raw_key = f"infra-{comp}-{key}"
            hypo_use, bs_use, extra_why = hypo, bs, ""
            # value-sharpening from the infra-md facts
            if comp == "route53" and key == "dns-ttl" and facts.get("dns_simple_cname"):
                scen = "Route 53 failover is MANUAL (simple CNAME, not health-based)"
                hypo_use = ("There is NO automatic DNS failover: the record is a simple "
                            "CNAME with no health check. Regional failover requires a "
                            "manual DNS change. Drill the manual runbook end-to-end and "
                            "measure real time-to-shift including TTL expiry.")
                bs_use = 5
                extra_why = " Evidence: 'Route53 (SIMPLE CNAME, not health-based)'."
            if comp == "nlb" and facts.get("nlb_idle_s"):
                extra_why = f" Real value: appnlb TCP idle {facts['nlb_idle_s']}s."
            if comp == "alb" and facts.get("alb_idle_s"):
                extra_why = f" Real value: appalb idle {facts['alb_idle_s']}s."
            if comp == "ecs-fargate" and key == "az" and facts.get("azs_perf"):
                extra_why = (f" Real values: PROD {facts.get('azs_prod')} AZs, PERF "
                             f"{facts.get('azs_perf')} AZs — see ENV_DELTA gap.")
            pkg.add(
                dedup_key=(DEDUP_ALIAS.get(raw_key, raw_key), nid), layer="Infra", target=nid,
                scenario=f"[{comp}] {scen}", hypothesis=hypo_use,
                why=f"Component: {comp} on {nid}. {source_note}{extra_why}",
                tooling=tool, base_l=bl, base_s=bs_use, criticality=crit,
                data_class=j_dc, journey=jid, status=status, wave=wave,
                assertion=f"[{jid}] steps through {nid} meet the journey SLO under this fault")

    # ---- evidence-triggered from the scan card (unchanged: these carry real citations
    #      and OVERRIDE the declared-status equivalents where both exist) ----
    if card:
        cfg = (card.get("config") or {})
        ka = cfg.get("kafka_auth") or {}
        if ka.get("protocol") == "SASL_SSL":
            pkg.add(
                dedup_key=("infra-kafka-mbus-kafka-unreachable", nid), layer="Infra", target=nid,
                scenario="[kafka-mbus] Kafka broker (MBUS) unreachable",
                hypothesis="Producers buffer via outbox without unbounded growth; "
                           "consumers idle without crash-loop; no-loss catch-up on restore.",
                why=f"Scan-card evidence: kafka_auth protocol={ka.get('protocol')}, "
                    f"mechanism={ka.get('mechanism')}, token_provider="
                    f"{ka.get('token_provider')} — this service is a confirmed MBUS client.",
                tooling="Network blackhole to MBUS (PERF)",
                base_l=3, base_s=5, criticality=crit, data_class=j_dc, journey=jid,
                status="Proposed (evidence-confirmed)",
                assertion=f"[{jid}] zero event loss for this journey across a broker outage")
            pkg.add(
                dedup_key=("infra-kafka-mbus-kafka-ida", nid), layer="Infra", target=nid,
                scenario="[kafka-mbus] IDP token failure (Kafka OAUTHBEARER re-auth)",
                hypothesis="Sessions survive to re-auth; failed re-auth backs off sanely; "
                           "signature distinct from broker-unreachable.",
                why=f"Scan-card evidence: mechanism={ka.get('mechanism')}, "
                    f"token_provider={ka.get('token_provider')}.",
                tooling="IDP path block (PERF)",
                base_l=2, base_s=4, criticality=crit, data_class=j_dc, journey=jid,
                status="Proposed (evidence-confirmed)",
                assertion=f"[{jid}] event flow resumes after an IDP outage without restarts")
        for ob in card.get("outbound", []):
            if ob.get("via") == "http" and "kms" in str(ob.get("resolved", "")).lower():
                pkg.add(
                    dedup_key=("infra-kms-kms-deny", nid), layer="Infra", target=nid,
                    scenario="[kms] KMS (VaultKMS) unavailability — envelope-encryption path",
                    hypothesis="Publishing fails closed with an explicit KMS error (no "
                               "silent drop, no plaintext); halts visibly, resumes on restore.",
                    why=f"Scan-card evidence: outbound to "
                        f"{(ob.get('target_raw') or {}).get('host')} (resolved: "
                        f"{ob.get('resolved')}); retry="
                        f"{(ob.get('retry') or {}).get('http_retry_max_attempts')} attempts "
                        f"with unset per-attempt timeouts.",
                    tooling="Toxiproxy / deny on the KMS path (PERF)",
                    base_l=2, base_s=5, criticality=crit, data_class=j_dc, journey=jid,
                    status="Proposed (evidence-confirmed)",
                    assertion=f"[{jid}] no event published unencrypted or silently dropped")
        hosts = [h.get("name", "") for h in (card.get("dnsreg") or {}).get("hosts", [])]
        ro_hosts = [h for h in hosts if "-ro" in h]
        if ro_hosts:
            pkg.add(
                dedup_key=("infra-ro-endpoint", nid), layer="Infra", target=nid,
                scenario="Read-only endpoint failure — reader/writer path split",
                hypothesis="Loss of the -ro endpoint degrades reads per design without "
                           "impacting writes.",
                why=f"Scan-card evidence (dnsreg): read-only hostname(s) {ro_hosts} "
                    f"alongside {[h for h in hosts if h not in ro_hosts]}.",
                tooling="DNS/endpoint block on the -ro host (PERF)",
                base_l=2, base_s=3, criticality=crit, data_class=j_dc, journey=jid,
                status="Proposed (evidence-confirmed)",
                assertion=f"[{jid}] read steps degrade per design; write steps unaffected")

    # ---- evidence-triggered from journey edges: encrypted producers ----
    if node_kafka_roles.get("producer_encrypted"):
        pkg.add(
            dedup_key=("infra-topic-encryption", nid), layer="Infra", target=nid,
            scenario="[kms] Topic-encryption key denial mid-publish",
            hypothesis="Encrypt failure on publish fails closed and is alarmed; outbox "
                       "retains the event; no plaintext fallback.",
            why=f"Journey-edge evidence: this node's Kafka producer(s) declare "
                f"encryption={node_kafka_roles['producer_encrypted']}.",
            tooling="KMS key-policy deny (PERF)",
            base_l=2, base_s=4, criticality=crit, data_class=j_dc, journey=jid,
            status="Proposed (evidence-confirmed)",
            assertion=f"[{jid}] no plaintext publish and no event loss under key denial")




class Package:
    def __init__(self):
        self.rows = {}          # dedup key -> row dict (journeys accumulate)
        self.assertions = []    # (scenario_id, journey, assertion)
        self.gaps = []
        self.seq = {}

    def next_id(self, layer):
        pref = {"Infra": "GEN-INF", "App": "GEN-APP", "Upstream": "GEN-UP",
                "Downstream": "GEN-DWN", "Saga": "GEN-SAGA", "Detection": "GEN-DET",
                "Governance": "GEN-GOV"}[layer]
        self.seq[pref] = self.seq.get(pref, 0) + 1
        return f"{pref}-{self.seq[pref]:02d}"

    def add(self, dedup_key, layer, target, scenario, hypothesis, why, tooling,
            base_l, base_s, criticality, data_class, journey, status="Proposed",
            l_mod=0, wave=None, assertion=None):
        if dedup_key in self.rows:
            row = self.rows[dedup_key]
            if journey not in row["journeys"]:
                row["journeys"].append(journey)
            # keep the WORST (highest) score across journeys
            l, s = score(base_l, base_s, criticality, data_class, l_mod)
            row["L"], row["S"] = max(row["L"], l), max(row["S"], s)
        else:
            l, s = score(base_l, base_s, criticality, data_class, l_mod)
            row = {"id": self.next_id(layer), "layer": layer, "target": target,
                   "scenario": scenario, "hypothesis": hypothesis, "why": why,
                   "tooling": tooling, "L": l, "S": s, "journeys": [journey],
                   "status": status, "wave": wave}
            self.rows[dedup_key] = row
        if assertion:
            self.assertions.append((self.rows[dedup_key]["id"], journey, assertion))
        return self.rows[dedup_key]["id"]

    def gap(self, journey, gtype, detail):
        self.gaps.append({"journey": journey, "type": gtype, "detail": detail})


def ev_str(edge, *keys):
    parts = []
    for k in keys:
        v = edge.get(k)
        if v:
            parts.append(f"{k}={v}")
    return "; ".join(parts)


def generate(journeys, cards, declared_components=None, infra_mds=None):
    declared_components = declared_components or {}
    infra_mds = infra_mds or {}
    pkg = Package()

    for path, jf in journeys:
        jid = jf["journey"]
        crit = jf.get("criticality")
        prof = jf.get("profile", {})
        j_dc = prof.get("data_class")

        # ---------------- NODE templates (Infra + App layers) ----------------
        # pre-compute per-node kafka roles from edges (for evidence-triggered infra)
        kafka_roles = {}
        for e in jf.get("edges", []):
            if e.get("via") in ("kafka", "rollback-kafka"):
                enc = (e.get("producer") or {}).get("encryption")
                if enc:
                    kafka_roles.setdefault(e.get("from"), {}).setdefault("producer_encrypted", enc)
        for node in jf.get("nodes", []):
            nid = node["id"]
            surv = node.get("survivability") or {}
            if not node.get("scanned"):
                pkg.gap(jid, "UNSCANNED_NODE",
                        f"{nid} — no scan evidence; node-level scenarios NOT generated "
                        f"(accuracy contract: no evidence, no scenario). External/upstream "
                        f"platform? Confirm ownership and whether it enters scan scope.")
                continue

            comp = surv.get("compute") or {}
            dbf = surv.get("db_failover") or {}
            regions = surv.get("regions") or {}

            # INFRA: single-replica exposure (fires only if evidence says so)
            if comp.get("replica_count") == 1:
                pkg.add(
                    dedup_key=("infra-single-replica", nid), layer="Infra", target=nid,
                    scenario="Single-task kill — N=1 replica exposure",
                    hypothesis="ECS replaces the task within SLO; but with replica_count=1 "
                               "(and autoscale_min=1), EVERY task recycle is a brief 100% "
                               "capacity loss for this service in-region. Measure the real "
                               "customer-visible window.",
                    why=f"Infra evidence (journey node survivability): replica_count=1, "
                        f"autoscale_min={comp.get('autoscale_min')}, autoscale_max="
                        f"{comp.get('autoscale_max')} — a deploy, AZ event, or OOM kill "
                        f"momentarily zeroes this service's capacity.",
                    tooling="FIS aws:ecs:stop-task",
                    base_l=4, base_s=3, criticality=crit, data_class=j_dc, journey=jid,
                    assertion=f"[{jid}] journey step(s) through {nid} complete within SLO "
                              f"during the replacement window, or fail with a retriable error")
            # INFRA: Aurora Global failover (fires only if global_cluster confirmed)
            if dbf.get("global_cluster"):
                pkg.add(
                    dedup_key=("infra-db-failover", nid), layer="Infra", target=nid,
                    scenario="Aurora writer failover (in-region) under load",
                    hypothesis="Connections re-establish within pool timeout budget; "
                               "journey-level error window bounded.",
                    why=f"Infra evidence: global_cluster=true, secondary_regions="
                        f"{dbf.get('secondary_regions')}, deletion_protection="
                        f"{dbf.get('deletion_protection')}, backup_retention_days="
                        f"{dbf.get('backup_retention_days')}.",
                    tooling="FIS / failover-db-cluster",
                    base_l=3, base_s=4, criticality=crit, data_class=j_dc, journey=jid,
                    assertion=f"[{jid}] no journey step through {nid} produces a duplicate "
                              f"or lost write across the failover")
                pkg.add(
                    dedup_key=("infra-region-failover", nid), layer="Infra", target=nid,
                    scenario=f"Regional failover drill ({regions.get('primary')} -> "
                             f"{','.join(regions.get('secondary', []) or ['?'])})",
                    hypothesis="Cross-region promotion + traffic shift meets RTO; "
                               "secondary replica_count is sufficient for shifted load.",
                    why=f"Infra evidence: primary={regions.get('primary')}, secondary="
                        f"{regions.get('secondary')}, secondary_replica_count="
                        f"{comp.get('secondary_replica_count')} (note: =1 means the DR "
                        f"region also starts at N=1 — cold-capacity risk).",
                    tooling="Drill (scheduled, stakeholder sign-off)",
                    base_l=1, base_s=5, criticality=crit, data_class=j_dc, journey=jid,
                    wave=4,
                    assertion=f"[{jid}] end-to-end journey completes in DR region")
            # Survivability grade yellow -> escalate as focus
            if node.get("survivability_grade") == "yellow":
                pkg.add(
                    dedup_key=("infra-yellow-grade", nid), layer="Infra", target=nid,
                    scenario="Survivability-gap validation (graded YELLOW)",
                    hypothesis="The specific infra attribute that caused the yellow grade "
                               "(e.g. autoscale_max unset) degrades gracefully under load "
                               "spike / node loss.",
                    why=f"Infra evidence: survivability_grade=yellow — e.g. autoscale_max="
                        f"{comp.get('autoscale_max')}. Fix or accept explicitly; this "
                        f"scenario validates actual behavior at the gap.",
                    tooling="Load + FIS", base_l=3, base_s=4, criticality=crit,
                    data_class=j_dc, journey=jid, status="NEEDS_ASSESSMENT",
                    assertion=f"[{jid}] steps through {nid} survive a 2x load step "
                              f"without autoscale headroom")

            # INFRA: component templates — declared-architecture baseline +
            # evidence-triggered (Kafka/IDP/KMS/reader-endpoint/topic-encryption)
            card_obj = cards[nid][1] if nid in cards else None
            infra_component_templates(pkg, nid, card_obj, jf, jid, crit, j_dc,
                                      kafka_roles.get(nid, {}), declared_components,
                                      infra_md=infra_mds.get(nid))

            # APP layer: only if we hold this node's scan card
            if nid in cards:
                _, card = cards[nid]
                res = card.get("resiliency") or {}
                if card.get("circuit_breaker") == "none":
                    pkg.add(
                        dedup_key=("app-no-breaker", nid), layer="App", target=nid,
                        scenario="Missing circuit breaker — gap finding + post-fix trip test",
                        hypothesis="(Post-fix) breaker opens at configured threshold and "
                                   "fallback engages. Until then this is a CONFIGURATION "
                                   "GAP, not a runnable experiment.",
                        why=f"App evidence (scan card): circuit_breaker=\"none\"; retry "
                            f"instances {res.get('instances')} run "
                            f"{(res.get('config') or {}).get('time_attempts')} attempts "
                            f"with NO breaker above them — retries without a breaker "
                            f"amplify a struggling dependency.",
                        tooling="Code fix first; then Toxiproxy trip test",
                        base_l=4, base_s=4, criticality=crit, data_class=j_dc,
                        journey=jid, status="BLOCKED — config gap",
                        assertion=f"[{jid}] once fixed: breaker-open on a dependency "
                                  f"fault keeps {nid}'s other endpoints healthy")
                for ob in card.get("outbound", []):
                    if ob.get("via") == "kafka" and (ob.get("producer") or {}).get("delivery_timeout") in (None, "unset"):
                        pkg.add(
                            dedup_key=("app-outbox-timeout", nid, ob.get("topic")),
                            layer="App", target=f"{nid} -> {ob.get('topic')}",
                            scenario="Outbox publish with unset delivery timeout — broker stall test",
                            hypothesis="With delivery_timeout unset, a broker stall causes "
                                       "bounded outbox backlog (poller retries), not "
                                       "unbounded memory/thread growth in the publisher.",
                            why=f"App evidence: producer pattern=transactional-outbox, "
                                f"delivery_timeout=unset on topic {ob.get('topic')} — "
                                f"library default applies (version-dependent; NOT assumed "
                                f"here, verify).",
                            tooling="Broker blackhole (PERF) + outbox depth observation",
                            base_l=3, base_s=3, criticality=crit, data_class=j_dc,
                            journey=jid,
                            assertion=f"[{jid}] events for this journey are delivered "
                                      f"exactly-once (effective) after broker recovery")
                for ib in card.get("inbound", []):
                    idem = ib.get("server_idempotency") or ""
                    m = re.search(r"retention\s+PT(\d+)M", idem)
                    if m:
                        mins = m.group(1)
                        pkg.add(
                            dedup_key=("app-idem-window", nid), layer="App", target=nid,
                            scenario=f"Duplicate request outside idempotency window (PT{mins}M)",
                            hypothesis=f"A replayed request-id {mins}+ minutes later is "
                                       f"treated as intentionally-new — confirm that is "
                                       f"designed behavior, not an accidental gap.",
                            why=f"App evidence: server_idempotency=\"{idem}\" on inbound "
                                f"endpoint(s) — the {mins}-minute retention is a real coded "
                                f"number; behavior just past it is untested.",
                            tooling="Replay harness", base_l=3, base_s=3,
                            criticality=crit, data_class=j_dc, journey=jid,
                            assertion=f"[{jid}] no duplicate side-effect from an "
                                      f"in-window replay; post-window behavior documented")

        # ---------------- EDGE templates (Upstream / Downstream / Saga) ----------------
        node_scanned = {n["id"]: n.get("scanned", False) for n in jf.get("nodes", [])}
        saga_edges = []
        for edge in jf.get("edges", []):
            via = edge.get("via")
            frm, to = edge.get("from"), edge.get("to")
            e_crit = edge.get("criticality", crit)
            e_dc = edge.get("data_class", j_dc)
            grade = edge.get("grade")
            intent = edge.get("intent", "")

            if str(via).startswith("rollback-"):
                saga_edges.append(edge)

            # CONFIRMED DEFECT rows from red grades — regression guards, not hypotheses
            if grade == "red":
                cs = edge.get("callee_side") or {}
                note = (cs.get("consumer") or {}).get("note", "") if isinstance(cs.get("consumer"), dict) else ""
                pkg.add(
                    dedup_key=("defect", frm, to), layer="Downstream",
                    target=f"{frm} -> {to} ({via})",
                    scenario="CONFIRMED DEFECT — regression-guard scenario (post-fix)",
                    hypothesis="(Post-fix) the failure mode that earned the red grade "
                               "can no longer silently lose/corrupt data.",
                    why=f"Journey evidence: edge grade=red. {note[:400] if note else intent}",
                    tooling="File defect NOW; chaos scenario runs only after fix",
                    base_l=5, base_s=5, criticality=e_crit, data_class=e_dc,
                    journey=jid, status="CONFIRMED_DEFECT — fix first", wave=1,
                    assertion=f"[{jid}] the red-graded failure mode is closed and "
                              f"guarded by a repeatable injection")
                continue

            if grade == "pending-sme-input":
                pkg.gap(jid, "PENDING_SME_EDGE",
                        f"{frm} -> {to} ({via}): grade pending — resolve via SME/scan "
                        f"before generating injection scenarios against it. Intent: {intent}")
                continue

            # DOWNSTREAM: caller-side evidence present -> dependency-failure scenarios
            if edge.get("caller_known") and (edge.get("timeout") or edge.get("retry")):
                t = edge.get("timeout") or {}
                pkg.add(
                    dedup_key=("down-latency", frm, to), layer="Downstream",
                    target=f"{frm} -> {to}",
                    scenario="Dependency slow-but-alive at the coded timeout boundary",
                    hypothesis=f"Latency just under/over connect={t.get('connect_ms')}ms / "
                               f"read={t.get('read_ms')}ms fires the timeout as configured; "
                               f"retry ({edge.get('retry')}) then fallback "
                               f"({edge.get('fallback')}) engage in that order.",
                    why=f"Edge evidence: {ev_str(edge, 'timeout', 'retry', 'fallback', 'idempotency')} "
                        f"— real coded values, not placeholders. Intent: {intent}",
                    tooling="Toxiproxy latency injection",
                    base_l=4, base_s=4, criticality=e_crit, data_class=e_dc, journey=jid,
                    assertion=f"[{jid}] step '{intent}' degrades per its coded "
                              f"fallback, not by hanging")
                if edge.get("fallback") and "none" in str(edge.get("fallback")).lower():
                    pkg.add(
                        dedup_key=("down-hard-dep", frm, to), layer="Downstream",
                        target=f"{frm} -> {to}",
                        scenario="Hard dependency (no fallback) — outage propagation test",
                        hypothesis="A sustained dependency outage fails the journey step "
                                   "FAST and VISIBLY (no retry pile-up), and recovery is "
                                   "immediate on restore.",
                        why=f"Edge evidence: fallback=\"{edge.get('fallback')}\" — the "
                            f"caller rethrows on failure; this step hard-fails when the "
                            f"dependency is down. Intent: {intent}",
                        tooling="Toxiproxy 5xx / connection refuse",
                        base_l=3, base_s=4, criticality=e_crit, data_class=e_dc,
                        journey=jid, l_mod=1,
                        assertion=f"[{jid}] the hard failure of '{intent}' is "
                                  f"correctly surfaced to the journey initiator")
                fb = str(edge.get("fallback", ""))
                if "FAIL-SAFE" in fb or "emptySet" in fb:
                    pkg.add(
                        dedup_key=("gov-fail-open", frm, to), layer="Governance",
                        target=f"{frm} -> {to}",
                        scenario="FAIL-OPEN behavior — human review required (not a chaos test)",
                        hypothesis="N/A — this is a business/compliance question: is "
                                   "failing open on this check an approved decision?",
                        why=f"Edge evidence: fallback=\"{fb[:200]}\" — on any exception the "
                            f"check is skipped and the operation proceeds. Intent: {intent}",
                        tooling="Compliance/SME review",
                        base_l=1, base_s=5, criticality=e_crit, data_class=e_dc,
                        journey=jid, status="NEEDS COMPLIANCE REVIEW")

            # UPSTREAM: callee-side kafka consumer evidence -> consumer scenarios
            cs = edge.get("callee_side") or {}
            if via == "kafka" and edge.get("callee_known") and node_scanned.get(to):
                consumer = cs.get("consumer") or {}
                pkg.add(
                    dedup_key=("up-poison", to), layer="Upstream", target=to,
                    scenario="Poison message on consumed topic",
                    hypothesis="A malformed event is parked/DLQ'd and the partition "
                               "CONTINUES — one bad event cannot halt processing.",
                    why=f"Edge evidence: consumer ack={consumer.get('ack')}, "
                        f"error_handler={consumer.get('error_handler')}, "
                        f"dlt={consumer.get('dlt')} on topic {(cs.get('source_raw') or {}).get('topic')}.",
                    tooling="Malformed event on PERF topic",
                    base_l=4, base_s=4, criticality=e_crit, data_class=e_dc, journey=jid,
                    assertion=f"[{jid}] this journey's events keep flowing past a "
                              f"poison message")
                pkg.add(
                    dedup_key=("up-duplicate", to), layer="Upstream", target=to,
                    scenario="Duplicate event delivery (at-least-once) idempotency",
                    hypothesis="Reprocessing an already-consumed event produces no "
                               "duplicate side-effect downstream.",
                    why=f"Kafka is at-least-once; consumer ack={consumer.get('ack')}. "
                        f"Idempotency must be proven per consumer, not assumed.",
                    tooling="Re-publish processed events (PERF)",
                    base_l=4, base_s=4, criticality=e_crit, data_class=e_dc, journey=jid,
                    assertion=f"[{jid}] exactly-once EFFECTS hold for this journey's "
                              f"event chain")

            # UPSTREAM: http inbound with idempotency evidence -> surge/replay
            if via == "http" and edge.get("callee_known") and node_scanned.get(to) \
               and (cs.get("server_idempotency") or (edge.get("caller_side") or {}).get("server_idempotency")):
                pkg.add(
                    dedup_key=("up-replay", to), layer="Upstream", target=to,
                    scenario="Upstream retry-storm replay against idempotency guard",
                    hypothesis="A burst of same-request-id retries (upstream client "
                               "retrying aggressively) is deduplicated server-side; "
                               "exactly one effect.",
                    why=f"Edge evidence: server_idempotency="
                        f"{cs.get('server_idempotency') or (edge.get('caller_side') or {}).get('server_idempotency')}.",
                    tooling="Replay harness (burst mode)",
                    base_l=3, base_s=4, criticality=e_crit, data_class=e_dc, journey=jid,
                    assertion=f"[{jid}] '{intent}' produces exactly one effect under "
                              f"a same-id retry burst")

        # SAGA templates: per journey with rollback edges
        if saga_edges:
            chain = " ; ".join(f"{e.get('step')}: {e.get('from')}->{e.get('to')}"
                               for e in saga_edges)
            pkg.add(
                dedup_key=("saga-step-fail", jid), layer="Saga",
                target=f"{jid} rollback chain ({len(saga_edges)} steps)",
                scenario="Compensation step failure mid-rollback",
                hypothesis="Failing one rollback step leaves NO silent orphaned state: "
                           "the saga retries to completion or raises an actionable "
                           "stuck-state alert.",
                why=f"Journey evidence: compensation chain {chain[:400]}. Each step has "
                    f"its own coded timeout/retry/fallback; nothing tests one FAILING "
                    f"mid-rollback.",
                tooling="Toxiproxy on one rollback target during a drill",
                base_l=2, base_s=5, criticality=crit, data_class=j_dc, journey=jid,
                wave=3,
                assertion=f"[{jid}] a failed compensation is retried or alarmed; "
                          f"no orphaned records across the chain")
            pkg.add(
                dedup_key=("saga-visibility", jid), layer="Saga",
                target=f"{jid} rollback chain",
                scenario="Partial-rollback operational visibility",
                hypothesis="An operator can see per-step compensation completion from "
                           "monitoring alone, without reading code.",
                why="Same evidence base; this is the detection half of the saga risk.",
                tooling="Drill + dashboard review",
                base_l=2, base_s=4, criticality=crit, data_class=j_dc, journey=jid,
                wave=3,
                assertion=f"[{jid}] per-step rollback state is visible and alertable")

        # Unresolved identities -> coverage gaps
        for u in jf.get("unresolved", []):
            pkg.gap(jid, "UNRESOLVED_IDENTITY",
                    f"{u.get('service')} ({u.get('direction')}, via {u.get('via')}) — "
                    f"target/source identity unresolved; scenarios not generated for it.")

    return pkg

# ---------------------------------------------------------------- xlsx output
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

ARIAL = "Arial"; NAVY = "1F3864"; AMBER = "FCE4D6"; LIGHT = "D9E2F3"; GREY = "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def _hdr(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def _body(ws, row, values, fill=None, bold_first=True):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=ARIAL, size=9, bold=(bold_first and i == 1))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)

def write_package(pkg, out_path, journeys, cards):
    wb = openpyxl.Workbook()

    rows = sorted(pkg.rows.values(), key=lambda r: -(r["L"] * r["S"]))
    for row in rows:
        p = row["L"] * row["S"]
        if row.get("wave") is None:
            if "DEFECT" in row["status"] or "BLOCKED" in row["status"]:
                row["wave"] = 1
            elif row["layer"] == "Governance":
                row["wave"] = "-"
            elif p >= 16:
                row["wave"] = 1
            elif p >= 9:
                row["wave"] = 2
            else:
                row["wave"] = 3

    # -------- 1. Executive Summary / Focus Areas --------
    ws = wb.active
    ws.title = "Executive Summary"
    ws.column_dimensions["A"].width = 130
    t = ws.cell(row=1, column=1, value="Chaos Engineering Package — Executive Summary")
    t.font = Font(name=ARIAL, size=14, bold=True, color=NAVY)
    ws.cell(row=2, column=1,
            value=f"Generated from {len(journeys)} journey file(s) and {len(cards)} service "
                  f"scan card(s). {len(rows)} scenarios generated; {len(pkg.gaps)} coverage "
                  f"gaps. Every scenario cites the evidence that generated it; no scenario "
                  f"exists without evidence (accuracy contract)."
            ).font = Font(name=ARIAL, size=10, italic=True)
    r = 4
    tt = ws.cell(row=r, column=1, value="Focus areas (ranked by summed priority of contributing scenarios):")
    tt.font = Font(name=ARIAL, size=11, bold=True, color=NAVY); r += 1
    focus = {}
    for row in rows:
        key = (row["target"].split(" -> ")[0].split(" (")[0], row["layer"])
        focus.setdefault(key, {"sum": 0, "n": 0, "worst": None})
        f = focus[key]
        p = row["L"] * row["S"]
        f["sum"] += p; f["n"] += 1
        if f["worst"] is None or p > f["worst"][0]:
            f["worst"] = (p, row["scenario"], row["status"])
    ranked_focus = sorted(focus.items(), key=lambda kv: -kv[1]["sum"])[:12]
    for (target, layer), f in ranked_focus:
        line = (f"•  {target}  [{layer}] — {f['n']} scenario(s), summed priority {f['sum']}. "
                f"Top: \"{f['worst'][1]}\" ({f['worst'][2]})")
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # -------- 2. Scenario Catalog --------
    ws = wb.create_sheet("Scenario Catalog")
    headers = ["ID", "Layer", "Target", "Scenario", "Hypothesis",
               "Why (evidence-cited)", "Tooling", "Journeys",
               "Likelihood (1-5)", "Severity (1-5)", "Priority",
               "Likelihood Tier", "Risk Rating", "Wave", "Status"]
    widths = [12, 11, 26, 30, 38, 46, 22, 22, 8, 8, 8, 14, 12, 6, 24]
    _hdr(ws, 1, headers, widths)
    r = 2
    for row in rows:
        fill = None
        if "DEFECT" in row["status"]:
            fill = AMBER
        elif "BLOCKED" in row["status"] or "NEEDS" in row["status"]:
            fill = GREY
        _body(ws, r, [row["id"], row["layer"], row["target"], row["scenario"],
                      row["hypothesis"], row["why"], row["tooling"],
                      ", ".join(row["journeys"]), row["L"], row["S"], None,
                      None, None, row["wave"], row["status"]], fill=fill)
        ws.cell(row=r, column=11, value=f"=I{r}*J{r}")
        ws.cell(row=r, column=12,
                value=(f'=IF(I{r}=5,"Almost Certain",IF(I{r}=4,"Likely",'
                       f'IF(I{r}=3,"Possible",IF(I{r}=2,"Unlikely","Rare"))))'))
        ws.cell(row=r, column=13,
                value=(f'=IF(K{r}>=16,"Critical",IF(K{r}>=9,"High",'
                       f'IF(K{r}>=4,"Medium","Opportunistic")))'))
        for col in (11, 12, 13):
            c = ws.cell(row=r, column=col)
            c.font = Font(name=ARIAL, size=9)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
        r += 1
    LAST = r - 1
    ws.conditional_formatting.add(
        f"K2:K{LAST}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       mid_type="percentile", mid_value=50, mid_color="FFD966",
                       end_type="max", end_color="C00000"))
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:O{LAST}"

    # -------- 3. Assertion Matrix --------
    ws = wb.create_sheet("Assertion Matrix")
    _hdr(ws, 1, ["Scenario ID", "Journey", "Journey-specific assertion (pass condition)"],
         [12, 24, 100])
    r = 2
    for sid, journey, assertion in pkg.assertions:
        _body(ws, r, [sid, journey, assertion], bold_first=True)
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{max(r-1,1)}"

    # -------- 4. Coverage Gaps --------
    ws = wb.create_sheet("Coverage Gaps")
    _hdr(ws, 1, ["Journey", "Type", "Detail"], [24, 22, 100])
    r = 2
    for g in pkg.gaps:
        _body(ws, r, [g["journey"], g["type"], g["detail"]])
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{max(r-1,1)}"

    wb.save(out_path)
    return len(rows), len(pkg.assertions), len(pkg.gaps)




# ---------------------------------------------------------------- html output
def write_html(pkg, out_path, journeys, cards):
    """Single-file interactive dashboard. No external assets — safe for locked-down VDIs.
    Data lives in a <script type="application/json"> tag (read via JSON.parse) and all
    click handling uses event delegation — no inline handlers, no quote-escaping traps."""
    def rating(p):
        return "Critical" if p >= 16 else "High" if p >= 9 else "Medium" if p >= 4 else "Opportunistic"
    def tier(l):
        return {5: "Almost Certain", 4: "Likely", 3: "Possible", 2: "Unlikely", 1: "Rare"}.get(l, "-")

    rows = sorted(pkg.rows.values(), key=lambda r: -(r["L"] * r["S"]))
    amap = {}
    for sid, j, a in pkg.assertions:
        amap.setdefault(sid, []).append({"journey": j, "assertion": a})
    data = []
    for r in rows:
        p = r["L"] * r["S"]
        data.append({
            "id": r["id"], "layer": r["layer"], "target": r["target"],
            "scenario": r["scenario"], "hypothesis": r["hypothesis"], "why": r["why"],
            "tooling": r["tooling"], "journeys": r["journeys"], "L": r["L"], "S": r["S"],
            "P": p, "tier": tier(r["L"]), "rating": rating(p),
            "wave": r.get("wave"), "status": r["status"],
            "assertions": amap.get(r["id"], []),
        })
    payload = {
        "meta": {"journey_files": len(journeys), "scan_cards": len(cards),
                 "scenarios": len(data), "gaps": len(pkg.gaps),
                 "assertions": len(pkg.assertions)},
        "scenarios": data,
        "gaps": pkg.gaps,
    }
    # Prevent '</script>' termination inside the JSON island: '</' -> '<\/'
    blob = json.dumps(payload).replace("</", "<\\/")

    css = """
:root{--bg:#F6F7F9;--panel:#FFFFFF;--ink:#191C22;--mut:#5B6472;--line:#DDE1E8;
--defect:#C0361D;--ok:#1E6E46;--decl:#9A6A14;--gap:#6B7280;--acc:#24449C;
--crit:#C0361D;--high:#C97A10;--med:#3F6FB5;--opp:#8B93A1}
*{box-sizing:border-box;margin:0}
body{background:var(--bg);color:var(--ink);font:14px/1.5 "IBM Plex Sans","Segoe UI",system-ui,sans-serif}
header{position:sticky;top:0;z-index:9;background:var(--panel);border-bottom:1px solid var(--line);padding:14px 22px 10px}
h1{font-size:17px;font-weight:600;letter-spacing:.2px}
h1 span{color:var(--mut);font-weight:400}
.kpis{display:flex;gap:18px;margin:6px 0 8px;color:var(--mut);font-size:12.5px}
.kpis b{color:var(--ink);font-weight:600}
#spectrum{display:flex;height:26px;gap:1px;cursor:pointer}
#spectrum div{flex:1;min-width:2px;border-radius:1px;opacity:.85;transition:transform .08s}
#spectrum div:hover{transform:scaleY(1.35);opacity:1}
.bar{display:flex;flex-wrap:wrap;gap:8px;padding:12px 22px;align-items:center}
.chip{border:1px solid var(--line);background:var(--panel);border-radius:14px;padding:3px 11px;
font-size:12.5px;cursor:pointer;color:var(--mut);user-select:none}
.chip.on{border-color:var(--acc);color:var(--acc);background:#EEF2FC;font-weight:600}
#q{border:1px solid var(--line);border-radius:8px;padding:6px 10px;font:inherit;width:230px;background:var(--panel)}
main{padding:0 22px 60px;max-width:1280px}
.card{background:var(--panel);border:1px solid var(--line);border-radius:8px;margin:10px 0;overflow:hidden}
.row{display:grid;grid-template-columns:86px 1fr 120px 60px 96px;gap:12px;padding:11px 16px;cursor:pointer;align-items:center}
.row:hover{background:#FAFBFD}
.sid{font:12px "IBM Plex Mono",ui-monospace,monospace;color:var(--mut)}
.sc b{font-weight:600}
.sc .tg{color:var(--mut);font-size:12.5px}
.pill{font-size:11px;font-weight:600;border-radius:4px;padding:2px 8px;text-align:center;letter-spacing:.3px}
.p-Critical{background:#FBEAE6;color:var(--crit)} .p-High{background:#FBF1E2;color:var(--high)}
.p-Medium{background:#E9F0FA;color:var(--med)} .p-Opportunistic{background:#EFF1F4;color:var(--opp)}
.pr{font:600 15px "IBM Plex Mono",monospace;text-align:center}
.st{font-size:11.5px;color:var(--mut)}
.st.defect{color:var(--defect);font-weight:600}.st.ok{color:var(--ok)}.st.decl{color:var(--decl)}
.det{display:none;border-top:1px solid var(--line);padding:14px 16px;background:#FBFBFC}
.card.open .det{display:block}
.det h4{font-size:11px;text-transform:uppercase;letter-spacing:.8px;color:var(--mut);margin:12px 0 4px}
.det h4:first-child{margin-top:0}
.why{font:12.5px/1.55 "IBM Plex Mono",ui-monospace,monospace;background:#F1F3F6;border-left:3px solid var(--acc);
padding:10px 12px;border-radius:0 6px 6px 0;white-space:pre-wrap}
.as{font-size:13px;margin:3px 0;padding-left:16px;position:relative}
.as:before{content:"\\2192";position:absolute;left:0;color:var(--acc)}
.tabs{display:flex;gap:4px;margin:18px 0 4px}
.tab{padding:7px 16px;border:1px solid var(--line);border-bottom:none;border-radius:8px 8px 0 0;
background:#EEF0F4;cursor:pointer;font-size:13px;color:var(--mut)}
.tab.on{background:var(--panel);color:var(--ink);font-weight:600}
#gaps{display:none}
.gap{border:1px solid var(--line);border-left:3px solid var(--gap);background:var(--panel);
border-radius:6px;padding:10px 14px;margin:8px 0;font-size:13px}
.gap b{font:600 11px "IBM Plex Mono",monospace;letter-spacing:.4px}
.gap.CONFLICTING_EVIDENCE{border-left-color:var(--defect)}
.gap.ENV_DELTA_PERF_VALIDITY{border-left-color:var(--decl)}
.empty{color:var(--mut);padding:30px;text-align:center}
@media(max-width:820px){.row{grid-template-columns:70px 1fr 70px}.row .pill,.row .st{display:none}}
"""

    js = """
var D = JSON.parse(document.getElementById("pkgdata").textContent);
function $(s){return document.querySelector(s)}
function esc(t){return String(t).replace(/&/g,"&amp;").replace(/</g,"&lt;")}
document.title = "Chaos Package - " + D.meta.scenarios + " scenarios";
$("#sub").textContent = "generated from " + D.meta.journey_files + " journey file(s), " + D.meta.scan_cards + " scan card(s)";
$("#kpis").innerHTML = "<span><b>" + D.meta.scenarios + "</b> scenarios</span><span><b>" +
  D.meta.assertions + "</b> journey assertions</span><span><b>" + D.meta.gaps + "</b> coverage gaps</span>";
var RC = {Critical:"var(--crit)", High:"var(--high)", Medium:"var(--med)", Opportunistic:"var(--opp)"};
$("#spectrum").innerHTML = D.scenarios.map(function(s,i){
  return '<div style="background:' + RC[s.rating] + '" data-i="' + i + '" title="' + s.id + ' P' + s.P + '"></div>';
}).join("");
$("#spectrum").addEventListener("click", function(e){
  var i = e.target.getAttribute("data-i");
  if (i === null) return;
  var el = document.getElementById("r" + i);
  if (!el) return;
  el.classList.add("open");
  el.scrollIntoView({behavior:"smooth", block:"center"});
});
var layers = [], statuses = [];
D.scenarios.forEach(function(s){
  if (layers.indexOf(s.layer) < 0) layers.push(s.layer);
  var st = s.status.split(" ")[0].replace(/[^A-Za-z_]/g, "");
  if (st && statuses.indexOf(st) < 0) statuses.push(st);
});
var F = {layer:{}, status:{}, q:""};
function chips(sel, items, key){
  $(sel).innerHTML = items.map(function(v){
    return '<span class="chip" data-k="' + key + '" data-v="' + esc(v) + '">' + esc(v) + '</span>';
  }).join(" ");
}
chips("#layerChips", layers, "layer");
chips("#statusChips", statuses, "status");
document.querySelectorAll(".chip").forEach(function(c){
  c.addEventListener("click", function(){
    var set = F[c.getAttribute("data-k")], v = c.getAttribute("data-v");
    if (set[v]) delete set[v]; else set[v] = true;
    c.classList.toggle("on");
    render();
  });
});
$("#q").addEventListener("input", function(e){ F.q = e.target.value.toLowerCase(); render(); });
function any(o){ for (var k in o) return true; return false; }
function stClass(st){
  if (st.indexOf("DEFECT") >= 0) return "defect";
  if (st.indexOf("evidence-confirmed") >= 0) return "ok";
  if (st.indexOf("DECLARED") >= 0 || st.indexOf("NEEDS") >= 0) return "decl";
  return "";
}
function render(){
  var out = D.scenarios.map(function(s, i){
    if (any(F.layer) && !F.layer[s.layer]) return "";
    if (any(F.status)) {
      var hit = false;
      for (var v in F.status) if (s.status.indexOf(v) >= 0) hit = true;
      if (!hit) return "";
    }
    if (F.q && (s.scenario + s.target + s.why + s.id).toLowerCase().indexOf(F.q) < 0) return "";
    var asrt = s.assertions.map(function(a){ return '<div class="as">' + esc(a.assertion) + '</div>'; }).join("");
    return '<div class="card" id="r' + i + '">' +
      '<div class="row">' +
        '<span class="sid">' + s.id + '<br>' + esc(s.layer) + '</span>' +
        '<span class="sc"><b>' + esc(s.scenario) + '</b><div class="tg">' + esc(s.target) +
          ' | journeys: ' + s.journeys.map(esc).join(", ") + '</div></span>' +
        '<span class="pill p-' + s.rating + '">' + s.rating + '</span>' +
        '<span class="pr">' + s.P + '</span>' +
        '<span class="st ' + stClass(s.status) + '">' + esc(s.status) + '</span>' +
      '</div>' +
      '<div class="det">' +
        '<h4>Why this scenario exists (evidence)</h4><div class="why">' + esc(s.why) + '</div>' +
        '<h4>Hypothesis</h4><div>' + esc(s.hypothesis) + '</div>' +
        '<h4>Tooling | L' + s.L + ' (' + s.tier + ') x S' + s.S + ' | Wave ' + s.wave + '</h4><div>' + esc(s.tooling) + '</div>' +
        '<h4>Journey assertions</h4>' + (asrt || '<div class="as">-</div>') +
      '</div></div>';
  }).join("");
  $("#cat").innerHTML = out || '<div class="empty">No scenarios match the current filters.</div>';
}
$("#cat").addEventListener("click", function(e){
  var row = e.target.closest(".row");
  if (row) row.parentNode.classList.toggle("open");
});
$("#gaps").innerHTML = D.gaps.map(function(g){
  return '<div class="gap ' + g.type + '"><b>' + g.type + '</b> | ' + esc(g.journey) +
         '<div>' + esc(g.detail) + '</div></div>';
}).join("") || '<div class="empty">No gaps.</div>';
document.querySelectorAll(".tab").forEach(function(t){
  t.addEventListener("click", function(){
    document.querySelectorAll(".tab").forEach(function(x){ x.classList.remove("on"); });
    t.classList.add("on");
    $("#cat").style.display = t.getAttribute("data-t") === "cat" ? "block" : "none";
    $("#gaps").style.display = t.getAttribute("data-t") === "gaps" ? "block" : "none";
  });
});
render();
"""

    html = ("<!DOCTYPE html>\n<html lang=\"en\"><head><meta charset=\"utf-8\">\n"
            "<meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">\n"
            "<title>Chaos Engineering Package</title>\n<style>" + css + "</style></head><body>\n"
            "<header>\n  <h1>Chaos Engineering Package <span id=\"sub\"></span></h1>\n"
            "  <div class=\"kpis\" id=\"kpis\"></div>\n"
            "  <div id=\"spectrum\" title=\"Every scenario, colored by risk rating - click to jump\"></div>\n"
            "</header>\n<div class=\"bar\">\n"
            "  <input id=\"q\" placeholder=\"Search scenario, target, evidence...\">\n"
            "  <span id=\"layerChips\"></span><span id=\"statusChips\"></span>\n</div>\n"
            "<main>\n  <div class=\"tabs\">\n"
            "    <div class=\"tab on\" data-t=\"cat\">Scenarios</div>\n"
            "    <div class=\"tab\" data-t=\"gaps\">Coverage gaps</div>\n  </div>\n"
            "  <div id=\"cat\"></div>\n  <div id=\"gaps\"></div>\n</main>\n"
            "<script type=\"application/json\" id=\"pkgdata\">" + blob + "</script>\n"
            "<script>" + js + "</script></body></html>\n")
    with open(out_path, "w") as f:
        f.write(html)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, nargs="+",
                    help="dir(s)/file(s) to content-sniff for journey JSONs")
    ap.add_argument("--scan-cards", required=True, nargs="+",
                    help="dir(s)/file(s) containing service scan-card JSONs")
    ap.add_argument("--infra-md", default=None, nargs="+",
                    help="dir(s)/file(s) with <app>-infra-components-by-env.md files "
                         "(preferred per-service infra declaration; supersedes "
                         "--components for services it covers)")
    ap.add_argument("--components", default=None,
                    help="optional components.json: per-service infra component lists "
                         "(hand-filled now; replaced by infra scan output later)")
    ap.add_argument("--out", required=True)
    ap.add_argument("--html", default=None,
                    help="also write a single-file interactive HTML dashboard here")
    args = ap.parse_args()

    journeys, _ = discover(args.root)
    _, cards = discover(args.scan_cards)
    if not journeys:
        raise SystemExit("No journey files found (need top-level journey/nodes/edges).")

    declared = {}
    if args.components:
        raw = load_json_safe(args.components) or {}
        declared = {k: v for k, v in raw.items()
                    if not k.startswith("_") and isinstance(v, list)}

    infra_mds = {}
    for src in (args.infra_md or []):
        paths = [src] if os.path.isfile(src) else glob.glob(os.path.join(src, "**", "*infra-components*"), recursive=True)
        for p in paths:
            parsed = parse_infra_md(p)
            if parsed:
                infra_mds[parsed["service"]] = parsed
    if infra_mds:
        print(f"infra-md files ingested for: {sorted(infra_mds)}")

    pkg = generate(journeys, cards, declared, infra_mds)
    n_rows, n_assert, n_gaps = write_package(pkg, args.out, journeys, cards)
    if args.html:
        write_html(pkg, args.html, journeys, cards)
        print(f"Dashboard written: {args.html}")
    print(f"Package written: {args.out}")
    print(f"  scenarios: {n_rows}  assertions: {n_assert}  coverage gaps: {n_gaps}")


if __name__ == "__main__":
    main()
