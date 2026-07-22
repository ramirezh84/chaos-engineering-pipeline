#!/usr/bin/env python3
"""
consolidate.py — merge scan-output/**/*.json into the chaos engineering workbook.

Usage:
    python3 consolidate.py --workbook Chaos_Engineering_Scenario_Catalog_v2.xlsx \
                            --scan-output ./scan-output \
                            --out Chaos_Engineering_Scenario_Catalog_v3_scanned.xlsx

Never overwrites the input workbook. Always writes to --out. Safe to re-run as more
scan-output/*/findings.json files land — it always reads whatever is currently on disk.
"""
import argparse, json, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"; NAVY = "1F3864"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MISSING = {"value": None, "status": "MISSING_FROM_SCAN_OUTPUT", "file": "", "line": "", "snippet": ""}

def ev(d, *path):
    """Safely walk a nested evidence dict; return a MISSING stub if anything along the path is
    absent. Never fabricates a value — a gap in the subagent's JSON is reported as a gap."""
    cur = d
    for p in path:
        if not isinstance(cur, dict) or p not in cur:
            return dict(MISSING)
        cur = cur[p]
    if isinstance(cur, dict) and "status" in cur:
        return cur
    return {"value": cur, "status": "UNVERIFIED_SHAPE", "file": "", "line": "", "snippet": ""}

def load_json(path):
    if not os.path.exists(path):
        return None
    with open(path) as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            print(f"WARNING: {path} is not valid JSON — skipping", file=sys.stderr)
            return None

def hdr(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def body(ws, row, values):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=ARIAL, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER

def evtxt(e):
    if e is None:
        return "MISSING_FROM_SCAN_OUTPUT"
    val = e.get("value")
    status = e.get("status", "?")
    file = e.get("file", "")
    line = e.get("line", "")
    loc = f" ({file}:{line})" if file else ""
    return f"{val if val is not None else '—'} [{status}]{loc}"

def ensure_parent_dir(path):
    d = os.path.dirname(os.path.abspath(path))
    if d and not os.path.isdir(d):
        os.makedirs(d, exist_ok=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--scan-output", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    registry = load_json(os.path.join(args.scan_output, "repo-registry.json")) or []

    inv_ws = wb["Service Inventory"]
    acr_to_service = {}
    for r in range(2, inv_ws.max_row + 1):
        service = inv_ws.cell(row=r, column=2).value
        acr = inv_ws.cell(row=r, column=3).value
        if acr:
            acr_to_service[acr.lower()] = service

    def service_name(acronym):
        return acr_to_service.get((acronym or "").lower(), f"UNKNOWN ({acronym})")

    # remove any previously-generated scan sheets so re-runs don't duplicate rows
    for name in ["Scan Provenance", "Resilience Truth Table", "Timeout Chain",
                 "Kafka Consumer Audit", "Idempotency Suspicion Ranking",
                 "Anti-Pattern Findings", "Profile Mismatch Register",
                 "Infra Component Register"]:
        if name in wb.sheetnames:
            del wb[name]

    # ---------------- Scan Provenance ----------------
    ws = wb.create_sheet("Scan Provenance")
    hdr(ws, 1, ["Repo", "Path", "Branch", "SHA (after pull)", "Pulled At (UTC)",
                "Pull Status", "Notes"], [26, 40, 12, 14, 20, 16, 30])
    prov = load_json(os.path.join(args.scan_output, "provenance.json")) or []
    r = 2
    for p in prov:
        body(ws, r, [p.get("repo"), p.get("path"), p.get("branch"), p.get("sha_after_pull"),
                     p.get("pulled_at_utc"), p.get("pull_status"), p.get("notes")])
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(r-1,1)}"

    # ---------------- Resilience Truth Table ----------------
    ws = wb.create_sheet("Resilience Truth Table")
    hdr(ws, 1, ["Service", "Repo", "Client / Method", "Breaker Name",
                "Failure Rate Threshold", "Slow-Call Rate Threshold", "Slow-Call Duration Threshold",
                "Wait Duration (Open)", "Half-Open Permits", "Has Fallback?",
                "Retry Max Attempts", "Retry Has Jitter?", "Bulkhead Max Concurrent",
                "Suggested Scenario Targets (heuristic — review)"],
        [16,20,22,16,20,20,22,16,14,12,12,12,16,26])
    r = 2
    for repo_entry in registry:
        if not repo_entry.get("app_present"):
            continue
        findings = load_json(os.path.join(args.scan_output, repo_entry["app_repo"], "findings.json"))
        if not findings:
            continue
        svc = service_name(repo_entry["acronym"])
        for rc in findings.get("resilience_config", []):
            suggestions = []
            if ev(rc, "slow_call_rate_threshold").get("status") in ("NOT_CONFIGURED", "NOT_FOUND"):
                suggestions.append("CE-APP-02")
            if ev(rc, "failure_rate_threshold").get("status") in ("NOT_CONFIGURED", "NOT_FOUND"):
                suggestions.append("CE-APP-01")
            if ev(rc, "has_fallback_method").get("value") is False:
                suggestions.append("CE-APP-01")
            if ev(rc, "retry", "has_jitter").get("status") in ("NOT_CONFIGURED", "NOT_FOUND"):
                suggestions.append("CE-APP-03")
            body(ws, r, [
                svc, repo_entry["app_repo"], rc.get("client_or_method"), rc.get("breaker_name"),
                evtxt(ev(rc, "failure_rate_threshold")),
                evtxt(ev(rc, "slow_call_rate_threshold")),
                evtxt(ev(rc, "slow_call_duration_threshold")),
                evtxt(ev(rc, "wait_duration_in_open_state")),
                evtxt(ev(rc, "permitted_calls_half_open")),
                evtxt(ev(rc, "has_fallback_method")),
                evtxt(ev(rc, "retry", "max_attempts")),
                evtxt(ev(rc, "retry", "has_jitter")),
                evtxt(ev(rc, "bulkhead", "max_concurrent_calls")),
                ", ".join(sorted(set(suggestions))) or "—",
            ])
            r += 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:N{max(r-1,1)}"

    # ---------------- Timeout Chain (merged app + infra) ----------------
    ws = wb.create_sheet("Timeout Chain")
    hdr(ws, 1, ["Service", "Repo", "Layer", "Key", "Value", "Status", "File", "Line"],
        [16, 20, 18, 30, 14, 16, 34, 8])
    r = 2
    for repo_entry in registry:
        acr = repo_entry["acronym"]
        svc = service_name(acr)
        if repo_entry.get("app_present"):
            findings = load_json(os.path.join(args.scan_output, repo_entry["app_repo"], "findings.json"))
            if findings:
                for t in findings.get("timeout_chain", []):
                    e = t.get("evidence", {})
                    body(ws, r, [svc, repo_entry["app_repo"], t.get("layer"), t.get("key"),
                                 e.get("value"), e.get("status"), e.get("file"), e.get("line")])
                    r += 1
        if repo_entry.get("infra_present"):
            infra = load_json(os.path.join(args.scan_output, repo_entry["infra_repo"], "infra_findings.json"))
            if infra:
                for res in infra.get("resources", []):
                    if res.get("component_category") == "Edge / ingress":
                        for attr_name, e in res.get("attributes", {}).items():
                            if attr_name in ("idle_timeout", "deregistration_delay",
                                             "integration_timeout", "throttle_burst_limit",
                                             "throttle_rate_limit"):
                                body(ws, r, [svc, repo_entry["infra_repo"],
                                             f"infra:{res.get('resource_type')}",
                                             attr_name, e.get("value"), e.get("status"),
                                             e.get("file"), e.get("line")])
                                r += 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:H{max(r-1,1)}"

    # ---------------- Kafka Consumer Audit ----------------
    ws = wb.create_sheet("Kafka Consumer Audit")
    hdr(ws, 1, ["Service", "Repo", "Listener Method", "Topic", "Group ID",
                "auto.offset.reset", "enable.auto.commit", "max.poll.interval.ms",
                "Error Handler?", "DLQ Configured?", "Suggested Scenario Targets"],
        [16, 20, 22, 20, 18, 16, 16, 16, 14, 14, 22])
    r = 2
    for repo_entry in registry:
        if not repo_entry.get("app_present"):
            continue
        findings = load_json(os.path.join(args.scan_output, repo_entry["app_repo"], "findings.json"))
        if not findings:
            continue
        svc = service_name(repo_entry["acronym"])
        for kc in findings.get("kafka_consumers", []):
            suggestions = []
            offset_reset = ev(kc, "auto_offset_reset")
            if str(offset_reset.get("value")).lower() == "latest":
                suggestions.append("CE-KFK-04")  # candidate silent-message-loss risk — human review
            if ev(kc, "error_handler_present").get("value") is False:
                suggestions.append("CE-KFK-04")
            if ev(kc, "dlq_configured").get("status") in ("NOT_CONFIGURED", "NOT_FOUND"):
                suggestions.append("CE-KFK-04")
                suggestions.append("CE-PIPE-05")
            body(ws, r, [
                svc, repo_entry["app_repo"], kc.get("listener_method"), kc.get("topic"),
                kc.get("group_id"), evtxt(offset_reset), evtxt(ev(kc, "enable_auto_commit")),
                evtxt(ev(kc, "max_poll_interval_ms")), evtxt(ev(kc, "error_handler_present")),
                evtxt(ev(kc, "dlq_configured")), ", ".join(sorted(set(suggestions))) or "—",
            ])
            r += 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:K{max(r-1,1)}"

    # ---------------- Idempotency Suspicion Ranking ----------------
    ws = wb.create_sheet("Idempotency Suspicion Ranking")
    hdr(ws, 1, ["Service", "Repo", "Handler / Writer", "Evidence Type", "Suspicion Level",
                "File", "Line", "Snippet"], [16, 20, 26, 18, 14, 30, 8, 34])
    r = 2
    rows_to_sort = []
    for repo_entry in registry:
        if not repo_entry.get("app_present"):
            continue
        findings = load_json(os.path.join(args.scan_output, repo_entry["app_repo"], "findings.json"))
        if not findings:
            continue
        svc = service_name(repo_entry["acronym"])
        for ie in findings.get("idempotency_evidence", []):
            level_rank = {"high": 0, "medium": 1, "low": 2}.get(ie.get("suspicion_level"), 3)
            rows_to_sort.append((level_rank, [svc, repo_entry["app_repo"],
                                               ie.get("handler_or_writer"), ie.get("evidence_type"),
                                               ie.get("suspicion_level"), ie.get("file"),
                                               ie.get("line"), ie.get("snippet")]))
    for _, row in sorted(rows_to_sort, key=lambda x: x[0]):
        body(ws, r, row)
        r += 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:H{max(r-1,1)}"

    # ---------------- Anti-Pattern Findings ----------------
    ws = wb.create_sheet("Anti-Pattern Findings")
    hdr(ws, 1, ["Service", "Repo", "Pattern", "Description", "Severity Hint", "File", "Line", "Snippet"],
        [16, 20, 20, 34, 12, 30, 8, 34])
    r = 2
    for repo_entry in registry:
        if not repo_entry.get("app_present"):
            continue
        findings = load_json(os.path.join(args.scan_output, repo_entry["app_repo"], "findings.json"))
        if not findings:
            continue
        svc = service_name(repo_entry["acronym"])
        for ap in findings.get("anti_patterns", []):
            body(ws, r, [svc, repo_entry["app_repo"], ap.get("pattern_id"), ap.get("description"),
                         ap.get("severity_hint"), ap.get("file"), ap.get("line"), ap.get("snippet")])
            r += 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:H{max(r-1,1)}"

    # ---------------- Profile Mismatch Register ----------------
    ws = wb.create_sheet("Profile Mismatch Register")
    hdr(ws, 1, ["Service", "Repo", "Config Key", "PERF Value", "PROD Value", "Classification", "Note"],
        [16, 20, 32, 22, 22, 14, 34])
    r = 2
    for repo_entry in registry:
        if not repo_entry.get("app_present"):
            continue
        diff = load_json(os.path.join(args.scan_output, repo_entry["app_repo"], "profile-diff.json"))
        if not diff:
            continue
        svc = service_name(repo_entry["acronym"])
        for d in diff.get("diffs", []):
            body(ws, r, [svc, repo_entry["app_repo"], d.get("key"),
                         evtxt(d.get("perf_value")), evtxt(d.get("prod_value")),
                         d.get("classification"), d.get("note")])
            r += 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:G{max(r-1,1)}"

    # ---------------- Infra Component Register ----------------
    ws = wb.create_sheet("Infra Component Register")
    hdr(ws, 1, ["Service", "Repo", "IaC Tool", "Component Category", "Resource Type",
                "Resource ID/Name", "Key Attributes (evidence)", "File"],
        [16, 20, 12, 20, 22, 22, 44, 30])
    r = 2
    for repo_entry in registry:
        if not repo_entry.get("infra_present"):
            continue
        infra = load_json(os.path.join(args.scan_output, repo_entry["infra_repo"], "infra_findings.json"))
        if not infra:
            continue
        svc = service_name(repo_entry["acronym"])
        for res in infra.get("resources", []):
            attr_txt = "; ".join(f"{k}={evtxt(v)}" for k, v in res.get("attributes", {}).items())
            body(ws, r, [svc, repo_entry["infra_repo"], infra.get("iac_tool"),
                         res.get("component_category"), res.get("resource_type"),
                         res.get("resource_id_or_name"), attr_txt, res.get("file")])
            r += 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:H{max(r-1,1)}"

    ensure_parent_dir(args.out)
    wb.save(args.out)
    print(f"Wrote {args.out}")
    print("Open this workbook once in Excel (or run it through recalc.py if LibreOffice is "
          "available) to refresh the live formulas in the original tabs (Likelihood Tier, "
          "Risk Rating, Wave Plan counts, Coverage Matrix, Priority Model distributions) before "
          "final handoff to the chaos team.")

if __name__ == "__main__":
    main()
